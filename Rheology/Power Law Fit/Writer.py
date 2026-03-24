from pathlib import Path
import re
import numpy as np
from openpyxl import Workbook, load_workbook
from typing import Dict, Any, Optional


class ResultWriter:
    def __init__(self, output_file: Path):
        self.output_file = Path(output_file)

    def _extract_number_tuple(self, sample_name: str) -> Optional[tuple]:
        m = re.search(r'(\d+(?:\.\d+)*)', sample_name)
        if m:
            return tuple(int(x) for x in m.group(1).split('.'))
        return None

    def _sort_key(self, name: str):
        nums = self._extract_number_tuple(name)
        if nums is not None:
            return (0, nums, name)
        else:
            return (1, (), name)

    def _read_existing(self) -> Dict[str, Dict[str, Any]]:
        existing: Dict[str, Dict[str, Any]] = {}
        if not self.output_file.exists():
            return existing

        wb = load_workbook(self.output_file, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return existing

        header = [str(h).strip() if h is not None else "" for h in rows[0]]
        try:
            idx_sample = header.index("Sample")
        except ValueError:
            idx_sample = 0

        def find_header(name, fallback):
            try:
                return header.index(name)
            except ValueError:
                return fallback

        idx_K = find_header("K", 1)
        idx_n = find_header("n", 2)
        idx_r2 = find_header("R²", 3)

        for row in rows[1:]:
            if not row or len(row) <= idx_sample:
                continue
            sample = row[idx_sample]
            if sample is None:
                continue
            sample = str(sample)
            existing[sample] = {
                "K": row[idx_K] if len(row) > idx_K else None,
                "n": row[idx_n] if len(row) > idx_n else None,
                "R²": row[idx_r2] if len(row) > idx_r2 else None,
            }
        return existing

    def write(self, results: Dict[str, Dict[str, Any]]):
        """
        Write/merge results to the output Excel file, sorted naturally.
        `results` : { sample_name: {"K":..., "n":..., "R²":...}, ... }
        New entries override existing ones.
        """
        existing = self._read_existing()
        merged = existing.copy()
        for sample, vals in results.items():
            merged[sample] = vals

        sorted_samples = sorted(merged.keys(), key=self._sort_key)

        # ---------- robust numeric conversion ----------
        def to_float_safe(v):
            """Convert many representations to float or return None."""
            if v is None:
                return None
            # accept numpy scalar and python numbers
            if isinstance(v, (int, float, np.floating, np.integer)):
                val = float(v)
                if np.isnan(val):
                    return None
                return val
            s = str(v).strip()
            if s == "" or s.lower() in ("none", "nan", "na", "n/a"):
                return None
            # allow comma decimals
            s = s.replace(",", ".")
            try:
                val = float(s)
                if np.isnan(val):
                    return None
                return val
            except Exception:
                return None

        # -------- build groups and keep raw sample info for diagnostics --------
        groups: Dict[str, Dict[str, list]] = {}
        groups_samples: Dict[str, list] = {}  # group_name -> list of (sample_name, raw_vals)

        for sample in sorted_samples:
            vals = merged[sample]
            nums = self._extract_number_tuple(sample)
            if nums is None:
                group_name = sample
            else:
                if len(nums) >= 2:
                    group_name = f"DOE {nums[0]}.{nums[1]}"
                else:
                    group_name = f"DOE {nums[0]}"

            groups.setdefault(group_name, {"K": [], "n": [], "R²": []})
            groups_samples.setdefault(group_name, []).append((sample, vals))

            k_val = to_float_safe(vals.get("K"))
            n_val = to_float_safe(vals.get("n"))
            r2_val = to_float_safe(vals.get("R²"))

            if k_val is not None:
                groups[group_name]["K"].append(k_val)
            if n_val is not None:
                groups[group_name]["n"].append(n_val)
            if r2_val is not None:
                groups[group_name]["R²"].append(r2_val)

        # --------------------- diagnostics / logging ---------------------
        for g, samples in groups_samples.items():
            if not any(groups[g][f] for f in ("K", "n", "R²")):
                print(f"[WARN] Group '{g}' has NO numeric values. Samples in that group (raw values):")
                for sname, raw in samples:
                    print(f"    {sname}: K={raw.get('K')!r}, n={raw.get('n')!r}, R²={raw.get('R²')!r}")
            else:
                missing = []
                for sname, raw in samples:
                    if raw.get("K") is None or raw.get("n") is None or raw.get("R²") is None:
                        missing.append((sname, raw))
                if missing:
                    print(f"[INFO] Group '{g}' has numeric values but some samples contain missing entries:")
                    for sname, raw in missing:
                        print(f"    {sname}: K={raw.get('K')!r}, n={raw.get('n')!r}, R²={raw.get('R²')!r}")

        # ---------- compute group mean/std ----------
        group_means: Dict[str, Dict[str, Optional[float]]] = {}
        group_stds: Dict[str, Dict[str, Optional[float]]] = {}

        for g, vals in groups.items():
            def safe_mean(lst):
                if not lst:
                    return None
                if len(lst) == 1:
                    return float(lst[0])
                return float(np.mean(np.array(lst, dtype=float)))

            def safe_std(lst):
                if len(lst) <= 1:
                    return None
                return float(np.std(np.array(lst, dtype=float), ddof=1))

            group_means[g] = {
                "K": safe_mean(vals["K"]),
                "n": safe_mean(vals["n"]),
                "R²": safe_mean(vals["R²"]),
            }
            group_stds[g] = {
                "K": safe_std(vals["K"]),
                "n": safe_std(vals["n"]),
                "R²": safe_std(vals["R²"]),
            }

        # ---------- write workbook ----------
        sorted_groups = sorted(groups.keys(), key=self._sort_key)
        self.output_file.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.append(
            ["Sample", "K", "n", "R²",
             "Group (mean)", "K_mean", "n_mean", "R²_mean",
             "Group (std)", "K_std", "n_std", "R²_std"]
        )

        max_rows = max(len(sorted_samples), len(sorted_groups))
        for i in range(max_rows):
            row = []
            # samples block
            if i < len(sorted_samples):
                sample = sorted_samples[i]
                vals = merged[sample]
                row.extend([
                    sample,
                    vals.get("K") if vals.get("K") is not None else None,
                    vals.get("n") if vals.get("n") is not None else None,
                    vals.get("R²") if vals.get("R²") is not None else None
                ])
            else:
                row.extend([None, None, None, None])

            # group means block
            if i < len(sorted_groups):
                g = sorted_groups[i]
                gm = group_means[g]
                row.extend([
                    g,
                    gm["K"] if gm["K"] is not None else None,
                    gm["n"] if gm["n"] is not None else None,
                    gm["R²"] if gm["R²"] is not None else None
                ])
            else:
                row.extend([None, None, None, None])

            # group stds block
            if i < len(sorted_groups):
                g = sorted_groups[i]
                gs = group_stds[g]
                row.extend([
                    g,
                    gs["K"] if gs["K"] is not None else None,
                    gs["n"] if gs["n"] is not None else None,
                    gs["R²"] if gs["R²"] is not None else None
                ])
            else:
                row.extend([None, None, None, None])

            ws.append(row)

        wb.save(self.output_file)
        print(f"💾 Results written to {self.output_file} (rows: {max_rows})")
